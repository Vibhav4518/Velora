import logging
import traceback
import uuid
import zipfile

from decimal import Decimal, InvalidOperation
from pathlib import Path

from cloudinary_storage.storage import MediaCloudinaryStorage
from openpyxl import load_workbook
from reportlab.pdfgen import canvas

from django.conf import settings
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.core.files import File
from django.core.files.storage import default_storage
from django.db import transaction
from django.db.models import Avg, Count, F, Q, Sum
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.template.loader import render_to_string
from django.urls import reverse
from django.utils import timezone
from django.utils.text import get_valid_filename
from django.views.decorators.http import require_POST

from .decorators import (
    has_role_permission,
    role_permission_required,
)

from .forms import (
    BlogCategoryForm,
    BlogPostForm,
    ContactMessageForm,
    DealForm,
    ProductReviewForm,
    ProfileForm,
    UserAddressForm,
)

from .models import (
    BlogCategory,
    BlogPost,
    Cart,
    CartItem,
    Category,
    ContactMessage,
    Deal,
    Order,
    OrderItem,
    Permission,
    Product,
    ProductImage,
    ProductReview,
    Role,
    User,
    UserAddress,
    Wishlist,
)


logger = logging.getLogger(__name__)

# ==========================================================
# PRODUCT IMAGE STORAGE HELPERS
# ==========================================================


def get_product_image_storage():
    if getattr(settings, "USE_CLOUDINARY", False):
        return MediaCloudinaryStorage()

    return default_storage


def save_product_uploaded_image(uploaded_file):
    if not uploaded_file:
        raise ValueError("No image file was received.")

    storage = get_product_image_storage()

    original_name = get_valid_filename(
        uploaded_file.name
    )

    if not original_name:
        raise ValueError(
            "The uploaded image has an invalid filename."
        )

    extension = Path(original_name).suffix.lower()

    if extension not in ALLOWED_IMAGE_EXTENSIONS:
        raise ValueError(
            "Only JPG, JPEG, PNG and WebP images are allowed."
        )

    unique_name = (
        f"products/"
        f"{uuid.uuid4().hex}_"
        f"{original_name}"
    )

    saved_name = storage.save(
        unique_name,
        uploaded_file
    )

    if not saved_name:
        raise ValueError(
            "Storage did not return a saved filename."
        )

    return saved_name

def get_product_card_queryset():
    """
    Common Product queryset for product cards used on:
    - Home
    - Shop
    - Deals
    - Related Products
    - Category results

    Adds dynamic average rating and total review count.
    """

    return (
        Product.objects
        .filter(is_active=True)
        .select_related("category")
        .prefetch_related("images")
        .annotate(
            average_rating=Avg(
                "reviews__rating",
                filter=Q(
                    reviews__is_approved=True
                )
            ),

            review_count=Count(
                "reviews",
                filter=Q(
                    reviews__is_approved=True
                ),
                distinct=True
            ),
        )
    )

def get_customers_queryset():
    """
    Common customer query.
    Use this everywhere: dashboard, customer page, reports, etc.
    """
    return (
        User.objects
        .filter(is_superuser=False)
        .exclude(is_staff=True)
        .select_related("role")
        .order_by("-date_joined")
    )


def get_admins_queryset():
    """
    Common admin users query.
    """
    return (
        User.objects
        .filter(is_staff=True)
        .select_related("role")
        .order_by("-date_joined")
    )


def get_products_queryset():
    """
    Common products query.
    """
    return Product.objects.select_related("category").prefetch_related("images")


def get_orders_queryset():
    """
    Common orders query.
    """
    return (
        Order.objects
        .select_related("user", "address")
        .prefetch_related("items")
        .order_by("-order_date")
    )


def get_admin_permissions_context(user):
    return {
        "can_view_dashboard": has_role_permission(user, "Can View Dashboard"),
        "can_manage_admins": has_role_permission(user, "Can Manage Admins"),
        "can_manage_roles": has_role_permission(user, "Can Manage Roles"),
        "can_manage_products": has_role_permission(user, "Can Manage Products"),
        "can_manage_categories": has_role_permission(user, "Can Manage Categories"),
        "can_manage_orders": has_role_permission(user, "Can Manage Orders"),
        "can_manage_customers": has_role_permission(user, "Can Manage Customers"),
        "can_manage_payments": has_role_permission(user, "Can Manage Payments"),
        "can_view_reports": has_role_permission(user, "Can View Reports"),
        "can_manage_blog": has_role_permission(user, "Can Manage Blog"),
        "can_manage_deals": has_role_permission(user, "Can Manage Deals"),
    }
    
# ==========================================================
# BULK PRODUCT IMPORT SETTINGS
# ==========================================================

ALLOWED_IMAGE_EXTENSIONS = {
    ".jpg",
    ".jpeg",
    ".png",
    ".webp",
}

# Keep batches small on Render Free.
MAX_BULK_PRODUCTS = 50

# Maximum Excel size: 2 MB.
MAX_EXCEL_SIZE = 2 * 1024 * 1024

# Maximum ZIP size: 10 MB.
MAX_ZIP_SIZE = 100 * 1024 * 1024


# ==========================================================
# BULK PRODUCT IMPORT HELPERS
# ==========================================================

def clean_excel_value(value):
    """
    Convert an Excel cell value to cleaned text.
    """

    if value is None:
        return ""

    return str(value).strip()


def normalize_image_filename(filename):
    """
    Return only the base filename.

    Examples:

    images/shirt.jpg -> shirt.jpg
    products\\shoe.png -> shoe.png
    """

    filename = clean_excel_value(
        filename
    )

    if not filename:
        return ""

    normalized_path = filename.replace(
        "\\",
        "/"
    )

    return normalized_path.split("/")[-1]


def build_zip_image_map(images_archive):
    """
    Create a case-insensitive image lookup.

    Example result:

    {
        "shirt.jpg": "images/shirt.jpg",
        "shoe.png": "shoe.png",
    }
    """

    image_map = {}

    for zip_name in images_archive.namelist():
        if zip_name.endswith("/"):
            continue

        normalized_name = normalize_image_filename(
            zip_name
        )

        extension = Path(
            normalized_name
        ).suffix.lower()

        if extension not in ALLOWED_IMAGE_EXTENSIONS:
            continue

        image_map[
            normalized_name.lower()
        ] = zip_name

    return image_map


def upload_zip_product_image(
    images_archive,
    zip_image_map,
    image_filename,
):
    """
    Open one image directly from the ZIP and stream it to
    Cloudinary without reading the complete ZIP into memory.
    """

    normalized_name = normalize_image_filename(
        image_filename
    )

    if not normalized_name:
        raise ValueError(
            "The image filename is empty."
        )

    zip_entry = zip_image_map.get(
        normalized_name.lower()
    )

    if not zip_entry:
        raise ValueError(
            f"Image '{normalized_name}' was not found "
            "inside the ZIP file."
        )

    extension = Path(
        normalized_name
    ).suffix.lower()

    if extension not in ALLOWED_IMAGE_EXTENSIONS:
        raise ValueError(
            f"Image '{normalized_name}' has an "
            "unsupported file format."
        )

    with images_archive.open(
        zip_entry,
        mode="r",
    ) as image_stream:
        uploaded_file = File(
            image_stream,
            name=normalized_name,
        )

        return save_product_uploaded_image(
            uploaded_file
        )


def parse_decimal(
    value,
    field_name,
    required=False,
):
    """
    Convert an Excel price cell to Decimal.
    """

    cleaned_value = clean_excel_value(
        value
    )

    if not cleaned_value:
        if required:
            raise ValueError(
                f"{field_name} is required."
            )

        return None

    try:
        number = Decimal(
            cleaned_value
        )

    except (
        InvalidOperation,
        TypeError,
        ValueError,
    ) as error:
        raise ValueError(
            f"Invalid {field_name}: "
            f"{cleaned_value}"
        ) from error

    if number < 0:
        raise ValueError(
            f"{field_name} cannot be negative."
        )

    return number


def parse_stock(value):
    """
    Convert an Excel stock cell to a non-negative integer.
    """

    cleaned_value = clean_excel_value(
        value
    )

    if not cleaned_value:
        return 0

    try:
        stock = int(
            float(cleaned_value)
        )

    except (
        TypeError,
        ValueError,
    ) as error:
        raise ValueError(
            f"Invalid stock value: "
            f"{cleaned_value}"
        ) from error

    if stock < 0:
        raise ValueError(
            "Stock cannot be negative."
        )

    return stock


def delete_saved_storage_files(
    saved_names
):
    """
    Remove Cloudinary/local files created during a failed row.
    """

    storage = get_product_image_storage()

    for saved_name in saved_names:
        if not saved_name:
            continue

        try:
            storage.delete(
                saved_name
            )

        except Exception:
            logger.exception(
                "Could not delete failed image upload: %s",
                saved_name,
            )

# ==========================================================
# PRODUCT REVIEW HELPERS
# ==========================================================

def user_has_purchased_product(
    user,
    product
):
    """
    Returns True when the customer has purchased this product.

    Change the allowed order statuses if your project uses
    different status names.
    """

    if not user.is_authenticated:
        return False

    return OrderItem.objects.filter(
        order__user=user,
        order__order_status__in=[
            "Confirmed",
            "Shipped",
            "Delivered",
        ],
        product=product,
    ).exists()


def get_product_review_summary(product):

    approved_reviews = (
        ProductReview.objects
        .filter(
            product=product,
            is_approved=True
        )
    )

    summary = approved_reviews.aggregate(
        average_rating=Avg("rating"),
        total_reviews=Count("id")
    )

    average_rating = (
        summary["average_rating"]
        or 0
    )

    total_reviews = (
        summary["total_reviews"]
        or 0
    )

    rating_counts = {
        rating: approved_reviews.filter(
            rating=rating
        ).count()
        for rating in range(1, 6)
    }

    rating_percentages = {}

    for rating in range(1, 6):

        count = rating_counts[rating]

        rating_percentages[rating] = (
            round(
                (
                    count /
                    total_reviews
                ) * 100
            )
            if total_reviews
            else 0
        )

    return {
        "average_rating": round(
            average_rating,
            1
        ),
        "total_reviews": total_reviews,
        "rating_counts": rating_counts,
        "rating_percentages": (
            rating_percentages
        ),
    }


def product_review_json(review):

    return {
        "id": str(review.id),

        "rating": review.rating,

        "title": review.title or "",

        "comment": review.comment,

        "username": (
            review.user.get_full_name()
            or review.user.username
        ),

        "user_id": str(
            review.user.id
        ),

        "profile_image": (
            review.user.profile_image.url
            if getattr(
                review.user,
                "profile_image",
                None
            )
            else ""
        ),

        "is_verified_purchase": (
            review.is_verified_purchase
        ),

        "created_at": timezone.localtime(
            review.created_at
        ).strftime(
            "%d %b %Y"
        ),
    }

# ==========================================================
# WEBSITE PAGES
# ==========================================================

def home(request):
    categories = (
        Category.objects
        .filter(is_active=True)
        .annotate(
            product_count=Count(
                "products",
                filter=Q(products__is_active=True)
            )
        )
        .order_by("category_name")[:8]
    )

    products = (
        Product.objects
        .filter(is_active=True)
        .select_related("category")
        .prefetch_related("images")
    )

    featured_products = products.order_by("-created_at")[:8]
    new_arrivals = products.order_by("-created_at")[:6]

    wishlist_product_ids = []

    if request.user.is_authenticated:
        wishlist_product_ids = list(
            Wishlist.objects
            .filter(user=request.user)
            .values_list("product_id", flat=True)
        )

    context = {
        "categories": categories,
        "featured_products": featured_products,
        "new_arrivals": new_arrivals,
        "wishlist_product_ids": wishlist_product_ids,
    }
    

    return render(request, "home.html", context)

def header_product_search(request):
    query = request.GET.get("q", "").strip()

    if len(query) < 2:
        return JsonResponse({
            "success": True,
            "products": [],
            "message": "Enter at least 2 characters."
        })

    products = (
        Product.objects
        .filter(is_active=True)
        .filter(
            Q(product_name__icontains=query)
            | Q(short_description__icontains=query)
            | Q(long_description__icontains=query)
            | Q(category__category_name__icontains=query)
            | Q(sku__icontains=query)
        )
        .select_related("category")
        .prefetch_related("images")
        .annotate(
            average_rating=Avg(
                "reviews__rating",
                filter=Q(reviews__is_approved=True)
            ),
            review_count=Count(
                "reviews",
                filter=Q(reviews__is_approved=True),
                distinct=True
            )
        )
        .distinct()[:8]
    )

    product_data = []

    for product in products:
        first_image = product.images.first()

        product_data.append({
    "id": str(product.id),
    "name": product.product_name,
    "category": (
        product.category.category_name
        if product.category
        else "Uncategorized"
    ),
    "description": product.short_description or "",
    "price": float(product.price),
    "current_price": float(product.current_price),
    "has_deal": bool(product.has_active_deal),
    "rating": round(product.average_rating or 0, 1),
    "review_count": product.review_count or 0,
    "image": (
        first_image.image.url
        if first_image and first_image.image
        else ""
    ),
    "url": reverse(
        "product_details",
        kwargs={"product_id": product.id}
    ),
})

    return JsonResponse({
        "success": True,
        "products": product_data,
        "count": len(product_data),
        "query": query,
    })

def categories(request):
    categories = (
        Category.objects
        .filter(is_active=True)
        .annotate(
            product_count=Count(
                "products",
                filter=Q(products__is_active=True)
            )
        )
        .order_by("category_name")
    )

    return render(request, "categories.html", {
        "categories": categories,
    })


# ==========================================================
# CUSTOMER DEALS PAGE
# ==========================================================

def deals(request):

    now = timezone.now()

    active_deals = (
        Deal.objects
        .filter(
            is_active=True,
            start_date__lte=now,
            end_date__gt=now
        )
        .prefetch_related(
            "products",
            "categories"
        )
        .order_by(
            "-priority",
            "-created_at"
        )
    )

    featured_deal = active_deals.first()

    deal_products = (
        get_product_card_queryset()
        .filter(
            Q(deals__in=active_deals)
            | Q(category__deals__in=active_deals)
        )
        .distinct()
        .order_by("-created_at")
    )

    deal_categories = (
        Category.objects
        .filter(
            products__in=deal_products
        )
        .distinct()
        .order_by("category_name")
    )

    wishlist_product_ids = []

    if request.user.is_authenticated:

        wishlist_product_ids = list(
            Wishlist.objects
            .filter(user=request.user)
            .values_list(
                "product_id",
                flat=True
            )
        )

    context = {
        "active_deals": active_deals,
        "featured_deal": featured_deal,
        "deal_products": deal_products,
        "deal_categories": deal_categories,
        "wishlist_product_ids": wishlist_product_ids,
    }

    return render(
        request,
        "deals.html",
        context
    )


# ==========================================================
# CUSTOMER BLOG PAGE
# ==========================================================

def blog(request):

    posts = get_published_blog_posts()

    featured_post = (
        posts
        .filter(is_featured=True)
        .first()
    )

    if featured_post is None:
        featured_post = posts.first()

    categories = (
        BlogCategory.objects
        .filter(is_active=True)
        .annotate(
            total_posts=Count(
                "posts",
                filter=Q(
                    posts__is_published=True,
                    posts__published_at__lte=timezone.now()
                )
            )
        )
        .order_by("category_name")
    )

    popular_posts = (
        get_published_blog_posts()
        .order_by(
            "-views_count",
            "-published_at"
        )[:5]
    )

    context = {
        "posts": posts,
        "featured_post": featured_post,
        "categories": categories,
        "popular_posts": popular_posts,
    }

    return render(
        request,
        "blog.html",
        context
    )



# ==========================================================
# PUBLIC BLOG POPUP DETAILS
# ==========================================================

def public_blog_popup_details(request, slug):

    post = get_object_or_404(
        BlogPost.objects.select_related(
            "category",
            "author",
        ),
        slug=slug,
        is_published=True,
        published_at__lte=timezone.now(),
    )

    # Increase views safely
    BlogPost.objects.filter(
        id=post.id
    ).update(
        views_count=F("views_count") + 1
    )

    post.refresh_from_db(
        fields=["views_count"]
    )

    related_posts = (
        BlogPost.objects
        .filter(
            is_published=True,
            published_at__lte=timezone.now(),
            category=post.category,
        )
        .exclude(id=post.id)
        .select_related(
            "category",
            "author",
        )
        .order_by(
            "-is_featured",
            "-published_at",
        )[:3]
    )

    author_name = "Velora Team"
    author_image = ""

    if post.author:
        author_name = (
            post.author.get_full_name()
            or post.author.username
        )

        if getattr(
            post.author,
            "profile_image",
            None
        ):
            author_image = (
                post.author.profile_image.url
            )

    return JsonResponse({
        "success": True,
        "status": "success",

        "post": {
            "id": str(post.id),

            "title": post.title,

            "slug": post.slug,

            "short_description": (
                post.short_description or ""
            ),

            "content": (
                post.content or ""
            ),

            "category_name": (
                post.category.category_name
                if post.category
                else "Uncategorized"
            ),

            "author_name": author_name,

            "author_image": author_image,

            "image": (
                post.image.url
                if post.image
                else ""
            ),

            "is_featured": (
                post.is_featured
            ),

            "published_display": (
                timezone.localtime(
                    post.published_at
                ).strftime(
                    "%d %b %Y, %I:%M %p"
                )
                if post.published_at
                else "Not published"
            ),

            "views_count": (
                post.views_count
            ),

            "public_url": reverse(
                "blog_details",
                kwargs={
                    "slug": post.slug,
                },
            ),
        },

        "related_posts": [
            {
                "title": related.title,

                "category_name": (
                    related.category.category_name
                    if related.category
                    else "Uncategorized"
                ),

                "image": (
                    related.image.url
                    if related.image
                    else ""
                ),

                "details_url": reverse(
                    "public_blog_popup_details",
                    kwargs={
                        "slug": related.slug,
                    },
                ),
            }
            for related in related_posts
        ],
    })

# ==========================================================
# CUSTOMER CONTACT PAGE
# ==========================================================

def contact(request):

    form = ContactMessageForm()

    if request.method == "POST":

        form = ContactMessageForm(
            request.POST
        )

        if not form.is_valid():

            errors = {
                field: [
                    str(error)
                    for error in field_errors
                ]
                for field, field_errors
                in form.errors.items()
            }

            if request.headers.get(
                "X-Requested-With"
            ) == "XMLHttpRequest":

                return JsonResponse({
                    "success": False,
                    "status": "error",
                    "message": (
                        "Please correct the form errors."
                    ),
                    "errors": errors,
                }, status=400)

            messages.error(
                request,
                "Please fill all required fields correctly."
            )

            return render(
                request,
                "contact.html",
                {
                    "contact_form": form,
                }
            )

        contact_message = form.save()

        if request.headers.get(
            "X-Requested-With"
        ) == "XMLHttpRequest":

            return JsonResponse({
                "success": True,
                "status": "success",
                "message": (
                    "Your message has been sent successfully."
                ),
                "contact_id": str(
                    contact_message.id
                ),
            })

        messages.success(
            request,
            "Your message has been sent successfully."
        )

        return redirect(
            "contact_page"
        )

    return render(
        request,
        "contact.html",
        {
            "contact_form": form,
        }
    )


def shop(request):
    products = (
        Product.objects
        .filter(is_active=True)
        .select_related("category")
        .prefetch_related("images")
    )

    categories = Category.objects.filter(is_active=True)

    category_id = request.GET.get("category")
    search_query = request.GET.get("q", "").strip()

    selected_category = None

    if category_id:
        selected_category = Category.objects.filter(
            id=category_id,
            is_active=True
        ).first()

        if selected_category:
            products = products.filter(category=selected_category)

    if search_query:
        products = products.filter(
            Q(product_name__icontains=search_query) |
            Q(short_description__icontains=search_query) |
            Q(category__category_name__icontains=search_query)
        )

    products = products.order_by("-created_at")

    return render(request, "shop.html", {
        "products": products,
        "categories": categories,
        "selected_category": selected_category,
        "search_query": search_query,
    })


def product_details(request, product_id):

    product = get_object_or_404(
        Product.objects
        .select_related("category")
        .prefetch_related(
            "images",
            "reviews__user",
        ),
        id=product_id,
        is_active=True
    )

    # ==========================================
    # Active Deal
    # ==========================================

    active_deal = (
        Deal.objects.filter(
            is_active=True,
            start_date__lte=timezone.now(),
            end_date__gte=timezone.now()
        )
        .filter(products=product)
        .first()
    )

    # ==========================================
    # Wishlist
    # ==========================================

    is_wishlist = False

    if request.user.is_authenticated:
        is_wishlist = Wishlist.objects.filter(
            user=request.user,
            product=product
        ).exists()

    # ==========================================
    # Reviews
    # ==========================================

    product_reviews = (
        ProductReview.objects
        .filter(
            product=product,
            is_approved=True
        )
        .select_related("user")
        .order_by("-created_at")
    )

    review_summary = product_reviews.aggregate(
        average_rating=Avg("rating"),
        total_reviews=Count("id")
    )

    average_rating = review_summary["average_rating"] or 0
    total_reviews = review_summary["total_reviews"] or 0

    rating_counts = {}

    rating_percentages = {}

    for star in range(1, 6):

        count = product_reviews.filter(
            rating=star
        ).count()

        rating_counts[star] = count

        rating_percentages[star] = (
            round((count / total_reviews) * 100)
            if total_reviews
            else 0
        )

    # ==========================================
    # User Review
    # ==========================================

    user_review = None

    if request.user.is_authenticated:

        user_review = ProductReview.objects.filter(
            product=product,
            user=request.user
        ).first()

    # ==========================================
    # Related Products
    # ==========================================

    related_products = (
        Product.objects.filter(
            category=product.category,
            is_active=True
        )
        .exclude(id=product.id)
        .prefetch_related("images")
        [:8]
    )

    verified_purchase = False
    can_review = False
    
    if request.user.is_authenticated:
    
        verified_purchase = user_has_purchased_product(
            request.user,
            product
        )
    
        can_review = (
            verified_purchase
            or user_review is not None
        )
    
    review_form = ProductReviewForm(
        instance=user_review
    )


    context = {

    "product": product,
    "active_deal": active_deal,
    "is_wishlist": is_wishlist,
    "product_reviews": product_reviews,
    "average_rating": round(average_rating, 1),
    "total_reviews": total_reviews,
    "rating_counts": rating_counts,
    "rating_percentages": rating_percentages,
    "user_review": user_review,
    "review_form": review_form,
    "verified_purchase": verified_purchase,
    "can_review": can_review,
    "related_products": related_products,
    }

    return render(
        request,
        "productDetails.html",
        context
    )
    
@login_required
@require_POST
def save_product_review(request, product_id):

    product = get_object_or_404(
        Product,
        id=product_id,
        is_active=True
    )

    existing_review = ProductReview.objects.filter(
        product=product,
        user=request.user
    ).first()

    verified_purchase = user_has_purchased_product(
        request.user,
        product
    )

    if not verified_purchase and existing_review is None:

        return JsonResponse({
            "success": False,
            "status": "error",
            "message": "You can review this product after purchasing it."
        }, status=403)

    form = ProductReviewForm(
        request.POST,
        instance=existing_review
    )

    if not form.is_valid():

        errors = {
            field: [
                str(error)
                for error in field_errors
            ]
            for field, field_errors in form.errors.items()
        }

        return JsonResponse({
            "success": False,
            "status": "error",
            "message": "Please correct the review form.",
            "errors": errors,
        }, status=400)

    review = form.save(commit=False)

    review.product = product
    review.user = request.user
    review.is_verified_purchase = verified_purchase

    review.save()

    return JsonResponse({
        "success": True,
        "status": "success",
        "message": (
            "Your review was updated successfully."
            if existing_review
            else "Your review was added successfully."
        )
    })


@login_required
@require_POST
def delete_product_review(request, review_id):

    review = get_object_or_404(
        ProductReview,
        id=review_id,
        user=request.user
    )

    review.delete()

    return JsonResponse({
        "success": True,
        "status": "success",
        "message": "Your review was deleted successfully."
    })


# ==========================================================
# AUTHENTICATION
# ==========================================================

def user_registration(request):
    if request.method == "POST":
        first_name = request.POST.get("first_name")
        last_name = request.POST.get("last_name")
        email = request.POST.get("email")
        password = request.POST.get("password")
        mobile_no = request.POST.get("mobile_no")
        profile_image = request.FILES.get("profile_image")

        if not first_name or not last_name or not email or not password:
            messages.error(request, "All fields are required.")
            return render(request, "register.html")

        if User.objects.filter(email=email).exists():
            messages.error(request, "Email already exists.")
            return render(request, "register.html")

        username = email.split("@")[0]

        if User.objects.filter(username=username).exists():
            messages.error(request, "Username already exists.")
            return render(request, "register.html")

        role = Role.objects.filter(role_name="User").first()

        user = User.objects.create_user(
            first_name=first_name,
            last_name=last_name,
            email=email,
            username=username,
            password=password,
            role=role,
            mobile_no=mobile_no,
            is_staff=False,
            is_active=True,
        )

        if profile_image:
            user.profile_image = profile_image
            user.save()

        messages.success(request, "Registration successful. Please login.")
        return redirect("login_user")

    return render(request, "register.html")


def user_login(request):
    if request.user.is_authenticated:
        if request.user.is_staff:
            return redirect("admin_dashboard")
        return redirect("home_page")

    if request.method == "POST":
        email = request.POST.get("email")
        password = request.POST.get("password")

        if not email or not password:
            messages.error(request, "Email and password are required.")
            return render(request, "login.html")

        user_obj = User.objects.filter(email=email).first()

        if not user_obj:
            messages.error(request, "Email does not exist.")
            return render(request, "login.html")

        user = authenticate(
            request,
            username=user_obj.username,
            password=password
        )

        if user is None:
            messages.error(request, "Invalid password.")
            return render(request, "login.html")

        login(request, user)
        messages.success(request, "Login successful.")

        if user.is_staff:
            return redirect("admin_dashboard")

        return redirect("home_page")

    return render(request, "login.html")


def user_logout(request):
    logout(request)
    messages.success(request, "Logout successful.")
    return redirect("home_page")


# ==========================================================
# WISHLIST
# ==========================================================

@login_required
def add_to_wishlist(request, product_id):

    product = get_object_or_404(Product, id=product_id)

    wishlist, created = Wishlist.objects.get_or_create(
        user=request.user,
        product=product
    )

    if created:
        return JsonResponse({
            "status": "added"
        })

    wishlist.delete()

    return JsonResponse({
        "status": "removed"
    })

@login_required
def wishlist_page(request):

    wishlist_items = Wishlist.objects.filter(
        user=request.user
    ).select_related(
        "product",
        "product__category"
    ).prefetch_related(
        "product__images"
    )

    total_items = wishlist_items.count()

    context = {
        "wishlist_items": wishlist_items,
        "total_items": total_items,
    }

    return render(
        request,
        "wishlist.html",
        context
    )
    
@login_required
def remove_from_wishlist(request, item_id):

    item = get_object_or_404(
        Wishlist,
        id=item_id,
        user=request.user
    )

    item.delete()

    messages.success(request, "Removed from wishlist.")

    return redirect("wishlist_page")

@login_required
def move_to_cart(request, item_id):

    wishlist = get_object_or_404(
        Wishlist,
        id=item_id,
        user=request.user
    )

    cart, created = Cart.objects.get_or_create(
        user=request.user
    )

    CartItem.objects.get_or_create(
        cart=cart,
        product=wishlist.product,
        defaults={
            "quantity": 1,
            "price": wishlist.product.get_final_price(),
        }
    )

    wishlist.delete()

    messages.success(request, "Moved to cart.")

    return redirect("wishlist_page")


@login_required
def clear_wishlist(request):

    Wishlist.objects.filter(
        user=request.user
    ).delete()

    messages.success(request, "Wishlist cleared.")

    return redirect("wishlist_page")

# ==========================================================
# PROFILE
# ==========================================================

def address_data(address):
    return {
        "id": str(address.id),
        "full_name": address.full_name,
        "mobile_no": address.mobile_no,
        "address_line1": address.address_line1,
        "address_line2": address.address_line2 or "",
        "city": address.city,
        "state_name": address.state_name,
        "country_name": address.country_name,
        "pincode": address.pincode,
        "is_default": address.is_default,
    }


@login_required
def my_profile(request):
    address_form = UserAddressForm()

    addresses = UserAddress.objects.filter(user=request.user).order_by("-is_default", "-id")
    default_address = addresses.filter(is_default=True).first()
    other_addresses = addresses.filter(is_default=False)

    orders = Order.objects.filter(user=request.user).order_by("-order_date")[:5]
    wishlist_items = Wishlist.objects.filter(user=request.user).select_related("product")[:4]

    return render(request, "myProfile.html", {
        "address_form": address_form,
        "default_address": default_address,
        "other_addresses": other_addresses,
        "orders": orders,
        "wishlist_items": wishlist_items,
    })


@login_required
def update_profile_ajax(request):
    if request.method != "POST":
        return JsonResponse({"status": "error", "message": "Invalid request."})

    user = request.user

    user.first_name = request.POST.get("first_name", "").strip()
    user.last_name = request.POST.get("last_name", "").strip()
    user.username = request.POST.get("username", "").strip()
    user.email = request.POST.get("email", "").strip()
    user.mobile_no = request.POST.get("mobile_no", "").strip()

    if request.FILES.get("profile_image"):
        user.profile_image = request.FILES.get("profile_image")

    user.save()

    return JsonResponse({
        "status": "success",
        "message": "Profile updated successfully.",
        "user": {
            "first_name": user.first_name,
            "last_name": user.last_name,
            "username": user.username,
            "email": user.email,
            "mobile_no": user.mobile_no or "",
            "profile_image": user.profile_image.url if user.profile_image else "",
        }
    })


@login_required
def add_address_ajax(request):
    if request.method != "POST":
        return JsonResponse({"status": "error", "message": "Invalid request."})

    form = UserAddressForm(request.POST)

    if not form.is_valid():
        return JsonResponse({
            "status": "error",
            "message": "Please fill all address fields correctly."
        })

    address = form.save(commit=False)
    address.user = request.user

    has_address = UserAddress.objects.filter(user=request.user).exists()

    if not has_address:
        address.is_default = True

    if address.is_default:
        UserAddress.objects.filter(user=request.user).update(is_default=False)

    address.save()

    return JsonResponse({
        "status": "success",
        "message": "Address added successfully.",
        "address": address_data(address),
    })


@login_required
def edit_address_ajax(request, address_id):
    if request.method != "POST":
        return JsonResponse({"status": "error", "message": "Invalid request."})

    address = get_object_or_404(UserAddress, id=address_id, user=request.user)
    form = UserAddressForm(request.POST, instance=address)

    if not form.is_valid():
        return JsonResponse({
            "status": "error",
            "message": "Please fill all address fields correctly."
        })

    updated_address = form.save(commit=False)

    if updated_address.is_default:
        UserAddress.objects.filter(user=request.user).exclude(id=address_id).update(is_default=False)

    updated_address.save()

    return JsonResponse({
        "status": "success",
        "message": "Address updated successfully.",
        "address": address_data(updated_address),
    })


@login_required
def set_default_address_ajax(request, address_id):
    address = get_object_or_404(UserAddress, id=address_id, user=request.user)

    UserAddress.objects.filter(user=request.user).update(is_default=False)
    address.is_default = True
    address.save()

    return JsonResponse({
        "status": "success",
        "message": "Default address updated.",
        "address": address_data(address),
    })


@login_required
def delete_address_ajax(request, address_id):
    address = get_object_or_404(UserAddress, id=address_id, user=request.user)

    was_default = address.is_default
    address.delete()

    new_default = None

    if was_default:
        new_default = UserAddress.objects.filter(user=request.user).first()

        if new_default:
            new_default.is_default = True
            new_default.save()

    return JsonResponse({
        "status": "success",
        "message": "Address deleted successfully.",
        "deleted_id": str(address_id),
        "new_default": address_data(new_default) if new_default else None,
    })


    
# ==========================================================
# CART PRICE HELPERS
# ==========================================================

def refresh_cart_prices(cart):
    """Refresh cart prices using the latest active deals."""
    updated_items = []

    if not cart:
        return updated_items

    for item in cart.items.select_related("product"):
        latest_price = item.product.get_final_price()

        if item.price != latest_price:
            item.price = latest_price
            item.save(update_fields=["price"])
            updated_items.append(item)

    return updated_items


# ==========================================================
# CART
# ==========================================================

@login_required
def cart_page(request):
    cart, created = Cart.objects.get_or_create(user=request.user)

    price_updated_items = refresh_cart_prices(cart)

    cart_items = (
        CartItem.objects
        .filter(cart=cart)
        .select_related("product", "product__category")
        .prefetch_related("product__images")
    )

    subtotal = Decimal("0.00")
    discount = Decimal("0.00")

    for item in cart_items:
        item.subtotal = item.price * item.quantity
        item.original_subtotal = item.product.price * item.quantity
        item.discount_amount = max(
            item.original_subtotal - item.subtotal,
            Decimal("0.00")
        )

        subtotal += item.subtotal
        discount += item.discount_amount

    shipping = Decimal("0.00")
    total = subtotal + shipping

    if price_updated_items:
        messages.info(
            request,
            "Some cart prices were updated because a deal started or expired."
        )

    context = {
        "cart": cart,
        "cart_items": cart_items,
        "subtotal": subtotal,
        "shipping": shipping,
        "discount": discount,
        "total": total,
    }

    return render(request, "cart.html", context)


@login_required
def add_to_cart(request, product_id):
    product = get_object_or_404(Product, id=product_id, is_active=True)

    if product.stock_quantity <= 0:
        return JsonResponse({
            "status": "error",
            "message": "This product is out of stock."
        }, status=400)

    cart, created = Cart.objects.get_or_create(user=request.user)
    final_price = product.get_final_price()

    cart_item, created = CartItem.objects.get_or_create(
        cart=cart,
        product=product,
        defaults={
            "quantity": 1,
            "price": final_price,
        }
    )

    if not created:
        if cart_item.quantity >= product.stock_quantity:
            return JsonResponse({
                "status": "error",
                "message": "You cannot add more than the available stock."
            }, status=400)

        cart_item.quantity += 1
        cart_item.price = final_price
        cart_item.save(update_fields=["quantity", "price"])

    return JsonResponse({
        "status": "success",
        "message": "Product added to cart.",
        "price": str(final_price),
        "quantity": cart_item.quantity,
        "cart_count": cart.items.count(),
    })


@login_required
def remove_cart(request, product_id):
    cart = get_object_or_404(Cart, user=request.user)

    cart_item = get_object_or_404(
        CartItem,
        cart=cart,
        product_id=product_id
    )

    cart_item.delete()

    messages.success(request, "Product removed from cart.")
    return redirect("cart")

# ==========================================================
# CHECKOUT
# ==========================================================

@login_required
def checkout(request):
    cart = Cart.objects.filter(user=request.user).first()

    price_updated_items = refresh_cart_prices(cart)

    cart_items = (
        CartItem.objects
        .filter(cart=cart)
        .select_related("product")
        if cart else []
    )

    subtotal = sum(
        (item.price * item.quantity for item in cart_items),
        Decimal("0.00")
    )

    deal_savings = sum(
        (
            max(item.product.price - item.price, Decimal("0.00"))
            * item.quantity
            for item in cart_items
        ),
        Decimal("0.00")
    )

    shipping = (
        Decimal("0.00")
        if subtotal >= Decimal("999.00") and subtotal > 0
        else Decimal("99.00")
    )

    total = subtotal + shipping if subtotal > 0 else Decimal("0.00")

    addresses = UserAddress.objects.filter(user=request.user)
    selected_address = addresses.filter(is_default=True).first()

    if price_updated_items:
        messages.info(
            request,
            "Some prices were updated because a deal started or expired."
        )

    return render(request, "checkout.html", {
        "cart_items": cart_items,
        "subtotal": subtotal,
        "deal_savings": deal_savings,
        "shipping": shipping,
        "total": total,
        "addresses": addresses,
        "selected_address": selected_address,
    })


@login_required
def checkout_set_address(request, address_id):
    address = get_object_or_404(UserAddress, id=address_id, user=request.user)

    UserAddress.objects.filter(user=request.user).update(is_default=False)
    address.is_default = True
    address.save()

    return JsonResponse({
        "status": "success",
        "message": "Delivery address updated.",
        "address": {
            "id": str(address.id),
            "full_name": address.full_name,
            "mobile_no": address.mobile_no,
            "address_line1": address.address_line1,
            "address_line2": address.address_line2 or "",
            "city": address.city,
            "state_name": address.state_name,
            "country_name": address.country_name,
            "pincode": address.pincode,
        }
    })
    
# ==========================================================
# ADMIN CONTACT MESSAGES
# ==========================================================

@role_permission_required("Can Manage Customers")
@login_required
def admin_contact_messages(request):

    query = request.GET.get(
        "q",
        ""
    ).strip()

    contact_messages = (
        ContactMessage.objects
        .all()
        .order_by("-created_at")
    )

    if query:
        contact_messages = (
            contact_messages.filter(
                Q(name__icontains=query)
                | Q(email__icontains=query)
                | Q(phone__icontains=query)
                | Q(subject__icontains=query)
                | Q(message__icontains=query)
            )
        )

    context = {
        "contact_messages": contact_messages,

        "total_messages": (
            ContactMessage.objects.count()
        ),

        "new_messages": (
            ContactMessage.objects.filter(
                status="New"
            ).count()
        ),

        "in_progress_messages": (
            ContactMessage.objects.filter(
                status="In Progress"
            ).count()
        ),

        "resolved_messages": (
            ContactMessage.objects.filter(
                status="Resolved"
            ).count()
        ),

        "unread_messages": (
            ContactMessage.objects.filter(
                is_read=False
            ).count()
        ),

        "search_query": query,
    }

    context.update(
        get_admin_permissions_context(
            request.user
        )
    )

    return render(
        request,
        "adminContactMessages.html",
        context
    )


# ==========================================================
# CONTACT MESSAGE DETAILS
# ==========================================================

@role_permission_required("Can Manage Customers")
@login_required
def admin_contact_message_details(
    request,
    message_id
):

    contact_message = get_object_or_404(
        ContactMessage,
        id=message_id
    )

    if not contact_message.is_read:

        contact_message.is_read = True

        contact_message.save(
            update_fields=[
                "is_read",
                "updated_at",
            ]
        )

    return JsonResponse({
        "success": True,

        "contact": {
            "id": str(contact_message.id),
            "name": contact_message.name,
            "email": contact_message.email,
            "phone": contact_message.phone or "",
            "subject": contact_message.subject,
            "message": contact_message.message,
            "status": contact_message.status,
            "is_read": contact_message.is_read,

            "created_at": timezone.localtime(
                contact_message.created_at
            ).strftime(
                "%d %b %Y, %I:%M %p"
            ),
        },
    })


# ==========================================================
# UPDATE CONTACT STATUS
# ==========================================================

@role_permission_required("Can Manage Customers")
@login_required
@require_POST
def admin_update_contact_status(
    request,
    message_id
):

    contact_message = get_object_or_404(
        ContactMessage,
        id=message_id
    )

    new_status = request.POST.get(
        "status",
        ""
    ).strip()

    valid_statuses = {
        choice[0]
        for choice
        in ContactMessage.STATUS_CHOICES
    }

    if new_status not in valid_statuses:

        return JsonResponse({
            "success": False,
            "status": "error",
            "message": "Invalid message status.",
        }, status=400)

    contact_message.status = new_status
    contact_message.is_read = True

    contact_message.save(
        update_fields=[
            "status",
            "is_read",
            "updated_at",
        ]
    )

    return JsonResponse({
        "success": True,
        "status": "success",
        "message": (
            "Contact message status updated."
        ),
        "contact_status": new_status,
        "message_id": str(contact_message.id),
    })


# ==========================================================
# DELETE CONTACT MESSAGE
# ==========================================================

@role_permission_required("Can Manage Customers")
@login_required
@require_POST
def admin_delete_contact_message(
    request,
    message_id
):

    contact_message = get_object_or_404(
        ContactMessage,
        id=message_id
    )

    deleted_id = str(
        contact_message.id
    )

    contact_message.delete()

    return JsonResponse({
        "success": True,
        "status": "success",
        "message": (
            "Contact message deleted successfully."
        ),
        "deleted_id": deleted_id,
    })

# ==========================================================
# BLOG REUSABLE HELPERS
# ==========================================================

def is_ajax_request(request):
    return (
        request.headers.get("X-Requested-With")
        == "XMLHttpRequest"
    )


def form_errors_json(form):
    return {
        field: [
            str(error)
            for error in field_errors
        ]
        for field, field_errors in form.errors.items()
    }


def get_blog_posts_queryset():
    return (
        BlogPost.objects
        .select_related(
            "category",
            "author",
        )
        .order_by(
            "-is_featured",
            "-published_at",
            "-created_at",
        )
    )


def blog_category_json(category):
    return {
        "id": str(category.id),
        "category_name": category.category_name,
        "description": category.description or "",
        "is_active": category.is_active,
        "posts_count": category.posts.count(),
    }


def blog_post_json(post):
    return {
        "id": str(post.id),
        "title": post.title,
        "slug": post.slug,

        "short_description": (
            post.short_description or ""
        ),

        "content": (
            post.content or ""
        ),

        "category_id": (
            str(post.category_id)
            if post.category_id
            else ""
        ),

        "category_name": (
            post.category.category_name
            if post.category
            else "Uncategorized"
        ),

        "author_name": (
            post.author.username
            if post.author
            else "Velora Team"
        ),

        "is_published": (
            post.is_published
        ),

        "is_featured": (
            post.is_featured
        ),

        "status_text": (
            post.status_text
        ),

        "views_count": (
            post.views_count
        ),

        "published_at": (
            timezone.localtime(
                post.published_at
            ).strftime(
                "%Y-%m-%dT%H:%M"
            )
            if post.published_at
            else ""
        ),

        "published_display": (
            timezone.localtime(
                post.published_at
            ).strftime(
                "%d %b %Y, %I:%M %p"
            )
            if post.published_at
            else "Not published"
        ),

        "image": (
            post.image.url
            if post.image
            else ""
        ),

        "public_url": (
            reverse(
                "blog_details",
                kwargs={
                    "slug": post.slug
                },
            )
            if (
                post.slug
                and post.is_published
            )
            else ""
        ),
    }

def blog_success_response(
    request,
    *,
    message,
    redirect_name="admin_blog_posts",
    payload=None,
    reload_required=True,
):
    payload = payload or {}

    if is_ajax_request(request):
        return JsonResponse({
            "success": True,
            "status": "success",
            "message": message,
            "reload_required": reload_required,
            **payload,
        })

    messages.success(
        request,
        message,
    )

    return redirect(
        redirect_name
    )


def blog_error_response(
    request,
    *,
    message,
    errors=None,
    status=400,
    redirect_name="admin_blog_posts",
):
    if is_ajax_request(request):
        return JsonResponse({
            "success": False,
            "status": "error",
            "message": message,
            "errors": errors or {},
        }, status=status)

    messages.error(
        request,
        message,
    )

    return redirect(
        redirect_name
    )


# ==========================================================
# PUBLIC BLOG PAGE
# ==========================================================

def blog_page(request):
    query = request.GET.get(
        "q",
        ""
    ).strip()

    category_id = request.GET.get(
        "category",
        ""
    ).strip()

    now = timezone.now()

    posts = (
        get_blog_posts_queryset()
        .filter(
            is_published=True,
            published_at__lte=now,
        )
    )

    if query:
        posts = posts.filter(
            Q(title__icontains=query)
            | Q(short_description__icontains=query)
            | Q(content__icontains=query)
            | Q(
                category__category_name__icontains=query
            )
            | Q(author__username__icontains=query)
        )

    if category_id:
        posts = posts.filter(
            category_id=category_id
        )

    categories = (
        BlogCategory.objects
        .filter(is_active=True)
        .annotate(
            published_posts_count=Count(
                "posts",
                filter=Q(
                    posts__is_published=True,
                    posts__published_at__lte=now,
                ),
            )
        )
        .order_by("category_name")
    )

    featured_posts = (
        posts
        .filter(is_featured=True)[:3]
    )

    popular_posts = (
        posts
        .order_by(
            "-views_count",
            "-published_at",
        )[:5]
    )

    context = {
        "posts": posts,
        "categories": categories,
        "featured_posts": featured_posts,
        "popular_posts": popular_posts,
        "search_query": query,
        "selected_category": category_id,
        "total_posts": posts.count(),
    }

    return render(
        request,
        "blog.html",
        context,
    )


# ==========================================================
# NORMAL FULL BLOG DETAILS PAGE
# ==========================================================

def blog_details(request, slug):

    post = get_object_or_404(
        BlogPost.objects.select_related(
            "category",
            "author",
        ),
        slug=slug,
        is_published=True,
        published_at__lte=timezone.now(),
    )

    BlogPost.objects.filter(
        id=post.id
    ).update(
        views_count=F("views_count") + 1
    )

    post.refresh_from_db(
        fields=["views_count"]
    )

    related_posts = (
        BlogPost.objects
        .filter(
            category=post.category,
            is_published=True,
            published_at__lte=timezone.now(),
        )
        .exclude(id=post.id)
        .select_related(
            "category",
            "author",
        )
        .order_by(
            "-is_featured",
            "-published_at",
        )[:4]
    )

    return render(
        request,
        "blogDetails.html",
        {
            "post": post,
            "related_posts": related_posts,
        },
    )

# ==========================================================
# ADMIN BLOG MANAGEMENT
# ==========================================================

@login_required
@role_permission_required(
    "Can Manage Blog"
)
def admin_blog_posts(request):
    query = request.GET.get(
        "q",
        ""
    ).strip()

    posts = get_blog_posts_queryset()

    if query:
        posts = posts.filter(
            Q(title__icontains=query)
            | Q(
                short_description__icontains=query
            )
            | Q(
                category__category_name__icontains=query
            )
            | Q(
                author__username__icontains=query
            )
        )

    categories = (
        BlogCategory.objects
        .annotate(
            total_posts=Count(
                "posts"
            )
        )
        .order_by(
            "category_name"
        )
    )

    now = timezone.now()

    context = {
        "posts": posts,
        "categories": categories,
        "blog_form": BlogPostForm(),
        "category_form": BlogCategoryForm(),
        "total_posts": BlogPost.objects.count(),
        "published_posts": (
            BlogPost.objects
            .filter(
                is_published=True,
                published_at__lte=now,
            )
            .count()
        ),
        "draft_posts": (
            BlogPost.objects
            .filter(
                is_published=False
            )
            .count()
        ),
        "featured_posts": (
            BlogPost.objects
            .filter(
                is_featured=True
            )
            .count()
        ),
        "total_blog_categories": (
            BlogCategory.objects.count()
        ),
        "search_query": query,
    }

    context.update(
        get_admin_permissions_context(
            request.user
        )
    )

    return render(
        request,
        "adminBlogPosts.html",
        context,
    )


# ==========================================================
# SAVE BLOG POST - ADD AND EDIT
# ==========================================================

@login_required
@role_permission_required(
    "Can Manage Blog"
)
@require_POST
def admin_save_blog_post(request):
    post_id = request.POST.get(
        "post_id",
        ""
    ).strip()

    post = None

    if post_id:
        post = get_object_or_404(
            BlogPost,
            id=post_id,
        )

    form = BlogPostForm(
        request.POST,
        request.FILES,
        instance=post,
    )

    if not form.is_valid():
        return blog_error_response(
            request,
            message=(
                "Please correct the blog form errors."
            ),
            errors=form_errors_json(
                form
            ),
        )

    saved_post = form.save(
        commit=False
    )

    if saved_post.author_id is None:
        saved_post.author = (
            request.user
        )

    if (
        saved_post.is_published
        and not saved_post.published_at
    ):
        saved_post.published_at = (
            timezone.now()
        )

    if not saved_post.is_published:
        saved_post.published_at = None

    saved_post.save()

    message = (
        "Blog post updated successfully."
        if post
        else "Blog post created successfully."
    )

    return blog_success_response(
        request,
        message=message,
        payload={
            "post": blog_post_json(
                saved_post
            ),
        },
    )


# ==========================================================
# BLOG POST DETAILS FOR EDIT MODAL
# ==========================================================

@login_required
@role_permission_required(
    "Can Manage Blog"
)
def admin_blog_post_details(
    request,
    post_id,
):
    post = get_object_or_404(
        BlogPost.objects.select_related(
            "category",
            "author",
        ),
        id=post_id,
    )

    return JsonResponse({
        "success": True,
        "status": "success",
        "post": blog_post_json(
            post
        ),
    })


# ==========================================================
# TOGGLE BLOG PUBLISH STATUS
# ==========================================================

@login_required
@role_permission_required(
    "Can Manage Blog"
)
@require_POST
def admin_toggle_blog_status(
    request,
    post_id,
):
    post = get_object_or_404(
        BlogPost,
        id=post_id,
    )

    post.is_published = (
        not post.is_published
    )

    if post.is_published:
        if not post.published_at:
            post.published_at = (
                timezone.now()
            )
    else:
        post.published_at = None

    post.save(
        update_fields=[
            "is_published",
            "published_at",
            "updated_at",
        ]
    )

    return JsonResponse({
        "success": True,
        "status": "success",

        "message": (
            "Blog post published successfully."
            if post.is_published
            else "Blog post moved to draft successfully."
        ),

        "post": blog_post_json(
            post
        ),
    })


# ==========================================================
# TOGGLE FEATURED BLOG
# ==========================================================


@login_required
@role_permission_required(
    "Can Manage Blog"
)
@require_POST
def admin_toggle_blog_feature(
    request,
    post_id,
):
    post = get_object_or_404(
        BlogPost,
        id=post_id,
    )

    post.is_featured = (
        not post.is_featured
    )

    post.save(
        update_fields=[
            "is_featured",
            "updated_at",
        ]
    )

    return JsonResponse({
        "success": True,
        "status": "success",

        "message": (
            "Blog post marked as featured."
            if post.is_featured
            else "Blog post removed from featured."
        ),

        "post": blog_post_json(
            post
        ),
    })


# ==========================================================
# DELETE BLOG POST
# ==========================================================

@login_required
@role_permission_required(
    "Can Manage Blog"
)
@require_POST
def admin_delete_blog_post(
    request,
    post_id,
):
    post = get_object_or_404(
        BlogPost,
        id=post_id,
    )

    deleted_id = str(
        post.id
    )

    post.delete()

    return JsonResponse({
        "success": True,
        "status": "success",
        "message": (
            "Blog post deleted successfully."
        ),
        "deleted_id": deleted_id,
    })


# ==========================================================
# SAVE BLOG CATEGORY - ADD AND EDIT
# ==========================================================

@login_required
@role_permission_required(
    "Can Manage Blog"
)
@require_POST
def admin_save_blog_category(
    request,
):
    category_id = request.POST.get(
        "category_id",
        ""
    ).strip()

    category = None

    if category_id:
        category = get_object_or_404(
            BlogCategory,
            id=category_id,
        )

    form = BlogCategoryForm(
        request.POST,
        instance=category,
    )

    if not form.is_valid():
        return blog_error_response(
            request,
            message=(
                "Please correct the category form errors."
            ),
            errors=form_errors_json(
                form
            ),
        )

    saved_category = (
        form.save()
    )

    message = (
        "Blog category updated successfully."
        if category
        else "Blog category created successfully."
    )

    return blog_success_response(
        request,
        message=message,
        payload={
            "category": (
                blog_category_json(
                    saved_category
                )
            ),
        },
    )


# ==========================================================
# BLOG CATEGORY DETAILS FOR EDIT MODAL
# ==========================================================

@login_required
@role_permission_required(
    "Can Manage Blog"
)
def admin_blog_category_details(
    request,
    category_id,
):
    category = get_object_or_404(
        BlogCategory,
        id=category_id,
    )

    return JsonResponse({
        "success": True,
        "status": "success",
        "category": (
            blog_category_json(
                category
            )
        ),
    })


# ==========================================================
# TOGGLE BLOG CATEGORY STATUS
# ==========================================================

@login_required
@role_permission_required(
    "Can Manage Blog"
)
@require_POST
def admin_toggle_blog_category_status(
    request,
    category_id,
):
    category = get_object_or_404(
        BlogCategory,
        id=category_id,
    )

    category.is_active = (
        not category.is_active
    )

    category.save(
        update_fields=[
            "is_active",
        ]
    )

    message = (
        "Blog category activated successfully."
        if category.is_active
        else "Blog category deactivated successfully."
    )

    return blog_success_response(
        request,
        message=message,
        payload={
            "category": (
                blog_category_json(
                    category
                )
            ),
        },
        reload_required=False,
    )


# ==========================================================
# DELETE BLOG CATEGORY
# ==========================================================

@login_required
@role_permission_required(
    "Can Manage Blog"
)
@require_POST
def admin_delete_blog_category(
    request,
    category_id,
):
    category = get_object_or_404(
        BlogCategory,
        id=category_id,
    )

    if category.posts.exists():
        return blog_error_response(
            request,
            message=(
                "Move or delete this category's posts first."
            ),
        )

    deleted_id = str(
        category.id
    )

    category.delete()

    return blog_success_response(
        request,
        message=(
            "Blog category deleted successfully."
        ),
        payload={
            "deleted_id": deleted_id,
        },
        reload_required=False,
    )
# ==========================================================
# ADMIN DASHBOARD
# ==========================================================

@login_required
def admin_dashboard(request):
    if not request.user.is_staff:
        messages.error(request, "You are not allowed to access admin panel.")
        return redirect("home_page")

    if not has_role_permission(request.user, "Can View Dashboard"):
        messages.error(request, "You do not have permission to access dashboard.")
        return redirect("home_page")

    customers = get_customers_queryset()
    products = get_products_queryset()
    orders = get_orders_queryset()

    total_sales = orders.aggregate(total=Sum("total_amount"))["total"] or 0

    context = {
        "total_sales": total_sales,
        "total_orders": orders.count(),
        "total_customers": customers.count(),
        "total_products": products.count(),
        "recent_orders": orders[:5],
    }

    context.update(get_admin_permissions_context(request.user))

    return render(request, "adminDashboard.html", context)


# ==========================================================
# DEFAULT ROLES AND PERMISSIONS
# ==========================================================

def create_default_permissions_roles():

    permission_names = [
        "Can View Dashboard",

        "Can Manage Admins",
        "Can Manage Roles",
        "Can Manage Customers",

        "Can Manage Products",
        "Can Delete Products",
        "Can Manage Categories",

        "Can Manage Orders",
        "Can View Orders",
        "Can Manage Payments",

        "Can View Reports",

        "Can Buy Products",
        "Can Manage Wishlist",
        "Can Manage Deals",
        "Can Manage Blog",
    ]

    permissions = {}

    # Create missing permissions
    for name in permission_names:

        permission, created = Permission.objects.get_or_create(
            permission_name=name
        )

        permissions[name] = permission


    # ======================================================
    # SUPER ADMIN
    # ======================================================

    super_admin, created = Role.objects.get_or_create(
        role_name="Super Admin",
        defaults={
            "is_default": True
        }
    )

    super_admin.is_default = True

    super_admin.save(
        update_fields=["is_default"]
    )

    super_admin.permissions.set(
        Permission.objects.all()
    )


    # ======================================================
    # ADMIN
    # ======================================================

    admin, created = Role.objects.get_or_create(
        role_name="Admin",
        defaults={
            "is_default": True
        }
    )

    admin.is_default = True

    admin.save(
        update_fields=["is_default"]
    )

    admin.permissions.set([
        permissions["Can View Dashboard"],

        permissions["Can Manage Customers"],

        permissions["Can Manage Products"],
        permissions["Can Delete Products"],
        permissions["Can Manage Categories"],

        permissions["Can Manage Orders"],
        permissions["Can View Orders"],
        permissions["Can Manage Payments"],

        permissions["Can View Reports"],

        permissions["Can Manage Deals"],
        permissions["Can Manage Blog"],
    ])


    # ======================================================
    # USER
    # ======================================================

    user, created = Role.objects.get_or_create(
        role_name="User",
        defaults={
            "is_default": True
        }
    )

    user.is_default = True

    user.save(
        update_fields=["is_default"]
    )

    user.permissions.set([
        permissions["Can Buy Products"],
        permissions["Can Manage Wishlist"],
        permissions["Can View Orders"],
    ])
    
# ==========================================================
# ADMIN ROLE MANAGEMENT
# ==========================================================

@role_permission_required("Can Manage Roles")
@login_required
def admin_role_list(request):

    create_default_permissions_roles()

    search_query = request.GET.get(
        "q",
        ""
    ).strip()

    roles = Role.objects.prefetch_related(
        "permissions"
    ).order_by(
        "-is_default",
        "role_name"
    )

    if search_query:
        roles = roles.filter(
            role_name__icontains=search_query
        )

    permissions = Permission.objects.all().order_by(
        "permission_name"
    )

    context = {
        "roles": roles,
        "permissions": permissions,

        "total_roles": Role.objects.count(),

        "default_roles": Role.objects.filter(
            is_default=True
        ).count(),

        "total_permissions": Permission.objects.count(),
    }

    return render(
        request,
        "adminRoleList.html",
        context
    )

@role_permission_required("Can Manage Roles")
def admin_edit_role(request, role_id):
    role = get_object_or_404(Role, id=role_id)

    if role.is_default:
        messages.error(request, "Default role cannot be edited.")
        return redirect("admin_role_list")

    if request.method == "POST":
        role_name = request.POST.get("role_name")

        if not role_name:
            messages.error(request, "Role name is required.")
            return redirect("admin_role_list")

        if Role.objects.filter(role_name__iexact=role_name).exclude(id=role.id).exists():
            messages.error(request, "Role already exists.")
            return redirect("admin_role_list")

        role.role_name = role_name
        role.save()

        messages.success(request, "Role updated successfully.")
        return redirect("admin_role_list")

    return redirect("admin_role_list")

@role_permission_required("Can Manage Roles")
@login_required
def admin_add_role(request):
    if request.method == "POST":
        role_name = request.POST.get("role_name", "").strip()

        if not role_name:
            messages.error(request, "Role name is required.")

        elif Role.objects.filter(role_name__iexact=role_name).exists():
            messages.error(request, "Role already exists.")

        else:
            Role.objects.create(role_name=role_name)
            messages.success(request, "Role created successfully.")

    return redirect("admin_role_list")

@role_permission_required("Can Manage Roles")
def admin_delete_role(request, role_id):
    role = get_object_or_404(Role, id=role_id)

    if role.is_default:
        messages.error(request, "Default roles cannot be deleted.")
        return redirect("admin_role_list")

    role.delete()
    messages.success(request, "Role deleted successfully.")
    return redirect("admin_role_list")


@login_required
def admin_role_permission(request, role_id):

    role = get_object_or_404(
        Role,
        id=role_id
    )

    permissions = Permission.objects.all().order_by(
        "permission_name"
    )

    if request.method == "POST":

        selected_permissions = request.POST.getlist(
            "permissions"
        )

        role.permissions.set(
            selected_permissions
        )

        if request.headers.get(
            "X-Requested-With"
        ) == "XMLHttpRequest":

            return JsonResponse({
                "status": "success",
                "success": True,
                "message": (
                    f"Permissions updated for "
                    f"{role.role_name}."
                ),
                "permission_count": (
                    role.permissions.count()
                ),
            })

        return redirect(
            "admin_role_list"
        )

    context = {
        "role": role,
        "permissions": permissions,
        "selected_permission_ids": list(
            role.permissions.values_list(
                "id",
                flat=True
            )
        ),
    }

    return render(
        request,
        "adminRolePermission.html",
        context
    )


# ==========================================================
# ADMIN USERS
# ==========================================================

@role_permission_required("Can Manage Admins")
def admin_users(request):
    query = request.GET.get("q", "")

    # Common Admin Query
    users = get_admins_queryset()

    if query:
        users = users.filter(
            Q(username__icontains=query) |
            Q(first_name__icontains=query) |
            Q(last_name__icontains=query) |
            Q(email__icontains=query) |
            Q(mobile_no__icontains=query) |
            Q(role__role_name__icontains=query)
        )

    roles = Role.objects.all().order_by("role_name")

    # Common Query for Dashboard Cards
    admins = get_admins_queryset()

    context = {
        "users": users,
        "roles": roles,

        "total_admins": admins.count(),
        "active_admins": admins.filter(is_active=True).count(),
        "inactive_admins": admins.filter(is_active=False).count(),
        "total_roles": Role.objects.count(),
    }

    # Sidebar Permissions
    context.update(get_admin_permissions_context(request.user))

    return render(request, "adminUserList.html", context)

@role_permission_required("Can Manage Admins")
def add_admin_user(request):
    roles = Role.objects.exclude(role_name="User")

    if request.method == "POST":
        first_name = request.POST.get("first_name")
        last_name = request.POST.get("last_name")
        username = request.POST.get("username")
        email = request.POST.get("email")
        mobile_no = request.POST.get("mobile_no")
        role_id = request.POST.get("role")
        password = request.POST.get("password")
        confirm_password = request.POST.get("confirm_password")
        is_active = request.POST.get("is_active") == "1"

        if password != confirm_password:
            messages.error(request, "Password and confirm password do not match.")
            return redirect("admin_users")

        if User.objects.filter(username=username).exists():
            messages.error(request, "Username already exists.")
            return redirect("admin_users")

        if User.objects.filter(email=email).exists():
            messages.error(request, "Email already exists.")
            return redirect("admin_users")

        role = get_object_or_404(Role, id=role_id)

        user = User.objects.create_user(
            username=username,
            email=email,
            password=password,
            first_name=first_name,
            last_name=last_name,
            mobile_no=mobile_no,
            role=role,
            is_active=is_active,
            is_staff=True,
        )

        if request.FILES.get("profile_image"):
            user.profile_image = request.FILES.get("profile_image")
            user.save()

        messages.success(request, "Admin user created successfully.")
        return redirect("admin_users")

    return render(request, "adminAddUser.html", {"roles": roles})


@login_required
def edit_admin_user(request, id):

    admin = get_object_or_404(User, id=id)

    roles = Role.objects.all()

    if request.method == "POST":

        admin.first_name = request.POST.get("first_name")
        admin.last_name = request.POST.get("last_name")
        admin.username = request.POST.get("username")
        admin.email = request.POST.get("email")
        admin.mobile_no = request.POST.get("mobile_no")

        role_id = request.POST.get("role")

        if role_id:
            admin.role = Role.objects.get(id=role_id)

        if request.FILES.get("profile_image"):
            admin.profile_image = request.FILES.get("profile_image")

        admin.save()

        messages.success(request, "Admin updated successfully.")

        return redirect("admin_users")

    context = {
        "admin": admin,
        "roles": roles,
    }

    return render(request, "adminEditUser.html", context)


@login_required
def change_admin_role(request, id):
    admin_user = get_object_or_404(User, id=id)
    roles = Role.objects.all()

    if request.method == "POST":
        role_id = request.POST.get("role")

        if role_id:
            admin_user.role = get_object_or_404(Role, id=role_id)
            admin_user.save()
            messages.success(request, "Role changed successfully.")
            return redirect("admin_users")

        messages.error(request, "Please select a role.")

    return render(request, "adminChangeRole.html", {
        "admin_user": admin_user,
        "roles": roles,
    })


@login_required
def reset_admin_password(request, id):
    admin_user = get_object_or_404(User, id=id)

    if request.method == "POST":
        password = request.POST.get("password")
        confirm_password = request.POST.get("confirm_password")

        if password != confirm_password:
            messages.error(request, "Passwords do not match.")
            return redirect("reset_admin_password", id=id)

        admin_user.set_password(password)
        admin_user.save()

        messages.success(request, "Password reset successfully.")
        return redirect("admin_users")

    return render(request, "adminResetPassword.html", {
        "admin_user": admin_user,
    })
    



@role_permission_required("Can Manage Admins")
def delete_admin_user(request, id):
    user = get_object_or_404(User, id=id, is_staff=True)

    if request.user.id == user.id:
        messages.error(request, "You cannot delete your own account.")
        return redirect("admin_users")

    user.delete()
    messages.success(request, "Admin user deleted successfully.")
    return redirect("admin_users")


# ==========================================================
# ADMIN PROFILE
# ==========================================================

@login_required
def admin_profile(request):
    if not request.user.is_staff:
        messages.error(request, "You are not allowed to access admin profile.")
        return redirect("home_page")

    if request.method == "POST":
        user = request.user

        user.first_name = request.POST.get("first_name")
        user.last_name = request.POST.get("last_name")
        user.username = request.POST.get("username")
        user.email = request.POST.get("email")
        user.mobile_no = request.POST.get("mobile_no")

        if request.FILES.get("profile_image"):
            user.profile_image = request.FILES.get("profile_image")

        user.save()

        messages.success(request, "Profile updated successfully.")
        return redirect("admin_profile")

    return render(request, "adminProfile.html")



# ==========================================================
# ADMIN CATALOG
# ==========================================================

@role_permission_required("Can Manage Categories")
@login_required
def add_category(request):

    if request.method == "POST":

        category_name = request.POST.get("category_name")
        description = request.POST.get("description")
        is_active = request.POST.get("is_active") == "True"
        image = request.FILES.get("category_image")

        if Category.objects.filter(category_name__iexact=category_name).exists():
            messages.error(request, "Category already exists.")
            return redirect("category_list")

        category = Category.objects.create(
            category_name=category_name,
            description=description,
            is_active=is_active,
            category_image=image
        )

        messages.success(request, "Category added successfully.")

    return redirect("category_list")

@role_permission_required("Can Manage Categories")
@login_required
def edit_category_admin(request, id):
    category = get_object_or_404(Category, id=id)

    if request.method == "POST":
        category.category_name = request.POST.get("category_name")
        category.description = request.POST.get("description")
        category.is_active = request.POST.get("is_active") == "True"

        if request.FILES.get("category_image"):
            category.category_image = request.FILES.get("category_image")

        category.save()
        messages.success(request, "Category updated successfully.")
        return redirect("category_list")

    return redirect("category_list")


@role_permission_required("Can Manage Categories")
@login_required
def delete_category_admin(request, id):
    category = get_object_or_404(Category, id=id)
    category.delete()

    messages.success(request, "Category deleted successfully.")
    return redirect("category_list")

@role_permission_required("Can Manage Categories")
def category_list(request):

    categories = Category.objects.annotate(
        total_products=Count("products")
    )

    context = {
        "categories": categories,
        "total_products": Product.objects.count(),
    }

    return render(request,"adminCategoryList.html",context)

@role_permission_required("Can Manage Products")
@login_required
def add_product(request):
    categories = (
        Category.objects
        .filter(is_active=True)
        .order_by("category_name")
    )

    if request.method != "POST":
        return render(
            request,
            "adminAddProducts.html",
            {
                "categories": categories,
            }
        )

    product_name = request.POST.get(
        "product_name",
        ""
    ).strip()

    short_description = request.POST.get(
        "short_description",
        ""
    ).strip()

    long_description = request.POST.get(
        "long_description",
        ""
    ).strip()

    price = request.POST.get(
        "price",
        ""
    ).strip()

    stock_quantity = request.POST.get(
        "stock_quantity",
        ""
    ).strip()

    category_id = request.POST.get(
        "category_id",
        ""
    ).strip()

    is_active = (
        request.POST.get("is_active")
        == "True"
    )

    uploaded_images = request.FILES.getlist(
        "product_images"
    )

    if (
        not product_name
        or not price
        or not stock_quantity
        or not category_id
    ):
        return JsonResponse({
            "success": False,
            "status": "error",
            "message": (
                "Please fill all required product fields."
            ),
        }, status=400)

    if not uploaded_images:
        return JsonResponse({
            "success": False,
            "status": "error",
            "message": (
                "Please select at least one product image."
            ),
        }, status=400)

    if Product.objects.filter(
        product_name__iexact=product_name
    ).exists():
        return JsonResponse({
            "success": False,
            "status": "error",
            "message": "Product already exists.",
        }, status=400)

    category = get_object_or_404(
        Category,
        id=category_id
    )

    product = Product.objects.create(
        product_name=product_name,
        short_description=short_description,
        long_description=long_description,
        price=price,
        stock_quantity=stock_quantity,
        category=category,
        is_active=is_active,
    )

    created_images = []

    try:
        for index, uploaded_file in enumerate(
            uploaded_images
        ):
            saved_name = save_product_uploaded_image(
                uploaded_file
            )

            product_image = ProductImage.objects.create(
                product=product,
                image=saved_name,
                is_primary=(index == 0),
            )

            created_images.append(
                product_image
            )

    except Exception as error:
        for product_image in created_images:
            try:
                product_image.image.delete(
                    save=False
                )
            except Exception:
                pass

        product.delete()

        return JsonResponse({
            "success": False,
            "status": "error",
            "message": (
                f"Image upload failed: {error}"
            ),
        }, status=500)

    product.refresh_from_db()

    return JsonResponse({
        "success": True,
        "status": "success",
        "message": "Product added successfully.",
        "reload_required": True,

        "product": {
            "id": str(product.id),
            "name": product.product_name,
            "category": category.category_name,
            "price": str(product.price),
            "stock": product.stock_quantity,

            "status": (
                "Active"
                if product.is_active
                else "Inactive"
            ),

            "image": product.primary_image_url,

            "uploaded_images": [
                image.image.url
                for image in created_images
            ],
        },
    })

@login_required
@role_permission_required("Can Manage Products")
def admin_products(request):
    query = request.GET.get(
        "q",
        ""
    ).strip()

    products = (
        Product.objects
        .select_related("category")
        .prefetch_related("images")
        .order_by("-created_at")
    )

    if query:
        products = products.filter(
            Q(product_name__icontains=query)
            | Q(category__category_name__icontains=query)
            | Q(short_description__icontains=query)
        )

    context = {
        "products": products,

        "categories": (
            Category.objects
            .filter(is_active=True)
            .order_by("category_name")
        ),

        "total_products": (
            Product.objects.count()
        ),

        "active_products": (
            Product.objects
            .filter(is_active=True)
            .count()
        ),

        "total_categories": (
            Category.objects.count()
        ),

        "search_query": query,
    }

    context.update(
        get_admin_permissions_context(
            request.user
        )
    )

    return render(
        request,
        "adminProducts.html",
        context
    )

@role_permission_required("Can Manage Products")
@login_required
def edit_product_admin(request, id):
    product = get_object_or_404(
        Product.objects
        .select_related("category")
        .prefetch_related("images"),
        id=id
    )

    if request.method != "POST":
        return redirect(
            "admin_products"
        )

    product_name = request.POST.get(
        "product_name",
        ""
    ).strip()

    price = request.POST.get(
        "price",
        ""
    ).strip()

    stock_quantity = request.POST.get(
        "stock_quantity",
        ""
    ).strip()

    category_id = request.POST.get(
        "category_id",
        ""
    ).strip()

    if (
        not product_name
        or not price
        or not stock_quantity
        or not category_id
    ):
        return JsonResponse({
            "success": False,
            "status": "error",
            "message": (
                "Please fill all required fields."
            ),
        }, status=400)

    duplicate_exists = (
        Product.objects
        .filter(
            product_name__iexact=product_name
        )
        .exclude(id=product.id)
        .exists()
    )

    if duplicate_exists:
        return JsonResponse({
            "success": False,
            "status": "error",
            "message": (
                "Another product already uses this name."
            ),
        }, status=400)

    category = get_object_or_404(
        Category,
        id=category_id
    )

    uploaded_images = request.FILES.getlist(
        "product_images"
    )

    # Update normal product information.
    product.product_name = product_name

    product.short_description = request.POST.get(
        "short_description",
        ""
    ).strip()

    product.long_description = request.POST.get(
        "long_description",
        ""
    ).strip()

    product.price = price
    product.stock_quantity = stock_quantity
    product.category = category

    product.is_active = (
        request.POST.get("is_active")
        == "True"
    )

    product.save()

    created_images = []

    if uploaded_images:
        try:
            # Upload every new physical file first.
            for index, uploaded_file in enumerate(
                uploaded_images
            ):
                saved_name = save_product_uploaded_image(
                    uploaded_file
                )

                new_image = ProductImage.objects.create(
                    product=product,
                    image=saved_name,
                    is_primary=(index == 0),
                )

                created_images.append(
                    new_image
                )

            new_image_ids = [
                image.id
                for image in created_images
            ]

            # Remove the old records only after all new
            # uploads succeed.
            old_images = (
                ProductImage.objects
                .filter(product=product)
                .exclude(id__in=new_image_ids)
            )

            for old_image in old_images:
                try:
                    old_image.image.delete(
                        save=False
                    )
                except Exception:
                    pass

            old_images.delete()

        except Exception as error:
            for new_image in created_images:
                try:
                    new_image.image.delete(
                        save=False
                    )
                except Exception:
                    pass

                new_image.delete()

            return JsonResponse({
                "success": False,
                "status": "error",
                "message": (
                    f"Image upload failed: {error}"
                ),
            }, status=500)

    product.refresh_from_db()

    return JsonResponse({
        "success": True,
        "status": "success",
        "message": "Product updated successfully.",
        "reload_required": True,

        "entity": {
            "type": "product",
            "id": str(product.id),

            "fields": {
                "name": product.product_name,
                "category": category.category_name,
                "price": f"₹{product.price}",
                "stock": product.stock_quantity,

                "status": (
                    "Active"
                    if product.is_active
                    else "Inactive"
                ),

                "image": product.primary_image_url,
            },
        },

        "received_file_count": len(
            uploaded_images
        ),

        "uploaded_images": [
            image.image.url
            for image in created_images
        ],
    })
        
        
        
@role_permission_required("Can Delete Products")
@login_required
def delete_product_admin(request, id):
    product = get_object_or_404(Product, id=id)
    product.delete()

    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
        return JsonResponse({
            "status": "success",
            "message": "Product deleted successfully."
        })

    messages.success(request, "Product deleted successfully.")
    return redirect("admin_products")

@login_required
def my_orders(request):
    orders = (
        Order.objects
        .filter(user=request.user)
        .order_by("-order_date")
    )

    return render(
        request,
        "myOrders.html",
        {
            "orders": orders,
        }
    )


@login_required
def my_order_details(request, id):
    order = get_object_or_404(
        Order.objects.prefetch_related(
            "items__product"
        ),
        id=id,
        user=request.user
    )

    return render(
        request,
        "myOrderDetails.html",
        {
            "order": order,
        }
    )
    
@login_required
def place_order(request):
    if request.method != "POST":
        return redirect("checkout")

    address_id = request.POST.get("address_id")
    payment_method = request.POST.get("payment_method", "COD")
    razorpay_payment_id = request.POST.get("razorpay_payment_id")

    if not address_id:
        messages.error(request, "Please select delivery address.")
        return redirect("checkout")

    address = get_object_or_404(
        UserAddress,
        id=address_id,
        user=request.user
    )

    cart = get_object_or_404(Cart, user=request.user)
    price_updated_items = refresh_cart_prices(cart)

    cart_items = (
        CartItem.objects
        .filter(cart=cart)
        .select_related("product")
    )

    if not cart_items.exists():
        messages.error(request, "Your cart is empty.")
        return redirect("cart")

    for item in cart_items:
        if item.quantity > item.product.stock_quantity:
            messages.error(
                request,
                f"Only {item.product.stock_quantity} unit(s) of "
                f"{item.product.product_name} are available."
            )
            return redirect("checkout")

    subtotal = sum(
        (item.price * item.quantity for item in cart_items),
        Decimal("0.00")
    )

    shipping = (
        Decimal("0.00")
        if subtotal >= Decimal("999.00") and subtotal > 0
        else Decimal("99.00")
    )

    total_amount = subtotal + shipping

    order = Order.objects.create(
        user=request.user,
        address=address,
        order_number=f"VELORA-{timezone.now().strftime('%Y%m%d%H%M%S%f')}",
        total_amount=total_amount,
        order_status="Pending",
        payment_status="Paid" if razorpay_payment_id else "Pending",
    )

    for item in cart_items:
        product = item.product
        active_deal = product.get_active_deal()
        original_price = product.price
        final_price = product.get_final_price()
        item_total = final_price * item.quantity

        OrderItem.objects.create(
            order=order,
            product=product,
            product_name=product.product_name,
            quantity=item.quantity,
            original_price=original_price,
            price=final_price,
            discount_amount=max(
                original_price - final_price,
                Decimal("0.00")
            ),
            total_price=item_total,
            deal=active_deal,
        )

        product.stock_quantity -= item.quantity
        product.save(update_fields=["stock_quantity"])

    cart_items.delete()

    if price_updated_items:
        messages.info(
            request,
            "Your cart prices were refreshed before placing the order."
        )

    messages.success(request, "Order placed successfully.")
    return redirect("my_order_details", id=order.id)


@login_required
def cancel_order(request, id):
    order = get_object_or_404(Order, id=id, user=request.user)

    if order.order_status in ["Delivered", "Cancelled"]:
        messages.error(request, "This order cannot be cancelled.")
        return redirect("my_orders")

    order.order_status = "Cancelled"
    order.save()

    messages.success(request, "Order cancelled successfully.")
    return redirect("my_orders")

@login_required
def download_invoice(request, id):
    order = get_object_or_404(
        Order.objects.prefetch_related("items"),
        id=id,
        user=request.user
    )

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="invoice-{order.order_number}.pdf"'

    p = canvas.Canvas(response)

    p.setFont("Helvetica-Bold", 20)
    p.drawString(50, 800, "Velora Invoice")

    p.setFont("Helvetica", 12)
    p.drawString(50, 760, f"Order No: {order.order_number}")
    p.drawString(50, 740, f"Date: {order.order_date.strftime('%d %b %Y')}")
    p.drawString(50, 720, f"Customer: {order.user.username}")
    p.drawString(50, 700, f"Status: {order.order_status}")
    p.drawString(50, 680, f"Payment: {order.payment_status}")

    y = 640
    p.setFont("Helvetica-Bold", 12)
    p.drawString(50, y, "Product")
    p.drawString(300, y, "Qty")
    p.drawString(380, y, "Price")
    p.drawString(470, y, "Total")

    y -= 25
    p.setFont("Helvetica", 11)

    for item in order.items.all():
        p.drawString(50, y, item.product_name[:30])
        p.drawString(300, y, str(item.quantity))
        p.drawString(380, y, f"Rs. {item.price}")
        p.drawString(470, y, f"Rs. {item.total_price}")
        y -= 22

    y -= 20
    p.setFont("Helvetica-Bold", 14)
    p.drawString(380, y, "Grand Total:")
    p.drawString(470, y, f"Rs. {order.total_amount}")

    p.showPage()
    p.save()

    return response


# ==========================================================
# ADMIN DEAL MANAGEMENT
# ==========================================================

@role_permission_required("Can Manage Products")
@login_required
def admin_deals(request):
    deals_qs = (
        Deal.objects
        .prefetch_related("products", "categories")
        .all()
    )

    query = request.GET.get("q", "").strip()

    if query:
        deals_qs = deals_qs.filter(
            Q(title__icontains=query) |
            Q(description__icontains=query) |
            Q(products__product_name__icontains=query) |
            Q(categories__category_name__icontains=query)
        ).distinct()

    deals_list = list(deals_qs)

    context = {
        "deals": deals_list,
        "deal_form": DealForm(),
        "search_query": query,
        "total_deals": len(deals_list),
        "active_deals": sum(1 for deal in deals_list if deal.status == "active"),
        "scheduled_deals": sum(1 for deal in deals_list if deal.status == "scheduled"),
        "expired_deals": sum(1 for deal in deals_list if deal.status == "expired"),
        "inactive_deals": sum(1 for deal in deals_list if deal.status == "inactive"),
    }

    context.update(get_admin_permissions_context(request.user))

    return render(request, "adminDeals.html", context)


@role_permission_required("Can Manage Products")
@login_required
@require_POST
def admin_save_deal(request):
    deal_id = request.POST.get("deal_id")
    deal = None

    if deal_id:
        deal = get_object_or_404(Deal, id=deal_id)

    form = DealForm(
        request.POST,
        request.FILES,
        instance=deal
    )

    if not form.is_valid():
        errors = {
            field: [str(message) for message in messages_list]
            for field, messages_list in form.errors.items()
        }

        return JsonResponse({
            "success": False,
            "status": "error",
            "message": "Please correct the form errors.",
            "errors": errors,
        }, status=400)

    saved_deal = form.save()

    return JsonResponse({
    "success": True,
    "status": "success",
    "message": (
        "Deal updated successfully."
        if deal else
        "Deal created successfully."
    ),
    "deal_id": str(saved_deal.id),
})


@role_permission_required("Can Manage Products")
@login_required
def admin_deal_details(request, deal_id):
    deal = get_object_or_404(Deal, id=deal_id)

    return JsonResponse({
        "success": True,
        "deal": {
            "id": str(deal.id),
            "title": deal.title,
            "description": deal.description or "",
            "discount_type": deal.discount_type,
            "discount_value": str(deal.discount_value),
            "products": [str(product.id) for product in deal.products.all()],
            "categories": [str(category.id) for category in deal.categories.all()],
            "start_date": timezone.localtime(deal.start_date).strftime("%Y-%m-%dT%H:%M"),
            "end_date": timezone.localtime(deal.end_date).strftime("%Y-%m-%dT%H:%M"),
            "priority": deal.priority,
            "is_active": deal.is_active,
            "banner_image": deal.banner_image.url if deal.banner_image else "",
        }
    })


@role_permission_required("Can Manage Products")
@login_required
@require_POST
def admin_toggle_deal(request, deal_id):
    deal = get_object_or_404(Deal, id=deal_id)
    deal.is_active = not deal.is_active
    deal.save(update_fields=["is_active", "updated_at"])

    return JsonResponse({
        "success": True,
        "status": "success",
        "message": (
            "Deal activated successfully."
            if deal.is_active else
            "Deal deactivated successfully."
        ),
        "is_active": deal.is_active,
        "deal_status": deal.status,
    })


@role_permission_required("Can Manage Products")
@login_required
@require_POST
def admin_delete_deal(request, deal_id):
    deal = get_object_or_404(Deal, id=deal_id)
    deal.delete()

    return JsonResponse({
        "success": True,
        "status": "success",
        "message": "Deal deleted successfully.",
        "deleted_id": str(deal_id),
    })


# ==========================================================
# ADMIN SALES
# ==========================================================

@role_permission_required("Can Manage Orders")
@login_required
def admin_orders(request):
    query = request.GET.get("q", "")

    orders = Order.objects.select_related("user").prefetch_related("items").order_by("-order_date")

    if query:
        orders = orders.filter(
            Q(order_number__icontains=query) |
            Q(user__username__icontains=query) |
            Q(user__email__icontains=query)
        )

    context = {
        "orders": orders,
        "total_orders": Order.objects.count(),
        "pending_orders": Order.objects.filter(order_status__iexact="Pending").count(),
        "completed_orders": Order.objects.filter(order_status__iexact="Completed").count(),
        "total_revenue": Order.objects.aggregate(total=Sum("total_amount"))["total"] or 0,
    }

    return render(request, "adminOrders.html", context)


@role_permission_required("Can Manage Orders")
@login_required
def update_order_status(request, id):
    order = get_object_or_404(Order, id=id)

    if request.method == "POST":
        status = request.POST.get("order_status")

        if status:
            order.order_status = status
            order.save()

        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
           return JsonResponse({
               "status": "success",
               "message": "Order status updated.",
               "entity": {
                   "type": "order",
                   "id": str(order.id),
                   "fields": {
                       "status": order.order_status,
                   }
               }
           })

        messages.success(request, "Order status updated.")
        return redirect("admin_orders")

    return redirect("admin_orders")


@role_permission_required("Can Manage Orders")
@login_required
def delete_order_admin(request, id):
    order = get_object_or_404(Order, id=id)
    order.delete()

    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
        return JsonResponse({
            "status": "success",
            "message": "Order deleted successfully."
        })

    messages.success(request, "Order deleted successfully.")
    return redirect("admin_orders")

@login_required
def admin_payments(request):
    query = request.GET.get("q", "")

    orders = Order.objects.all().order_by("-order_date")

    if query:
        orders = orders.filter(
            Q(order_number__icontains=query) |
            Q(user__username__icontains=query) |
            Q(user__email__icontains=query)
        )

    context = {
        "orders": orders,
        "total_payments": Order.objects.count(),
        "paid_payments": Order.objects.filter(payment_status__iexact="Paid").count(),
        "pending_payments": Order.objects.filter(payment_status__iexact="Pending").count(),
        "total_amount": Order.objects.aggregate(total=Sum("total_amount"))["total"] or 0,
    }

    return render(request, "adminPayments.html", context)

# ==========================================================
# ADMIN CUSTOMERS
# ==========================================================

@role_permission_required("Can Manage Customers")
def admin_customers(request):
    query = request.GET.get("q", "")

    customers = get_customers_queryset()

    if query:
        customers = customers.filter(
            Q(username__icontains=query) |
            Q(first_name__icontains=query) |
            Q(last_name__icontains=query) |
            Q(email__icontains=query) |
            Q(mobile_no__icontains=query)
        )

    context = {
        "customers": customers,
        "total_customers": get_customers_queryset().count(),
        "active_customers": get_customers_queryset().filter(is_active=True).count(),
        "inactive_customers": get_customers_queryset().filter(is_active=False).count(),
    }

    context.update(get_admin_permissions_context(request.user))

    return render(request, "adminCustomers.html", context)

@role_permission_required("Can Manage Customers")
def edit_customer(request, id):
    customer = get_object_or_404(User, id=id)

    if request.method == "POST":
        customer.first_name = request.POST.get("first_name")
        customer.last_name = request.POST.get("last_name")
        customer.username = request.POST.get("username")
        customer.email = request.POST.get("email")
        customer.mobile_no = request.POST.get("mobile_no")

        customer.is_active = request.POST.get("is_active") == "True"

        if request.FILES.get("profile_image"):
            customer.profile_image = request.FILES.get("profile_image")

        customer.save()
        messages.success(request, "Customer updated successfully.")
        return redirect("admin_customers")

    return redirect("admin_customers")


@role_permission_required("Can Manage Customers")
def delete_customer(request, id):
    customer = get_object_or_404(User, id=id)
    customer.delete()

    messages.success(request, "Customer deleted successfully.")
    return redirect("admin_customers")



#-------------------------------------------------------------------



        
# ==========================================================
# BULK PRODUCT IMPORT PROCESSOR
# ==========================================================

def process_product_bulk_import(request):
    """
    Process one Excel file and one ZIP of product images.

    Excel columns, in this exact order:

    1. Category
    2. Product Name
    3. Price
    4. Offer Price
    5. Stock
    6. Short Description
    7. Description
    8. Image1
    9. Image2
    10. Image3
    """

    excel_file = request.FILES.get(
        "excel_file"
    )

    images_zip = request.FILES.get(
        "images_zip"
    )

    # ======================================================
    # REQUIRED FILES
    # ======================================================

    if not excel_file:
        return JsonResponse(
            {
                "success": False,
                "status": "error",
                "message": (
                    "Please select an Excel file."
                ),
            },
            status=400,
        )

    if not images_zip:
        return JsonResponse(
            {
                "success": False,
                "status": "error",
                "message": (
                    "Please select an images ZIP file."
                ),
            },
            status=400,
        )

    # ======================================================
    # FILE EXTENSIONS
    # ======================================================

    if not excel_file.name.lower().endswith(
        ".xlsx"
    ):
        return JsonResponse(
            {
                "success": False,
                "status": "error",
                "message": (
                    "Only XLSX Excel files are supported."
                ),
            },
            status=400,
        )

    if not images_zip.name.lower().endswith(
        ".zip"
    ):
        return JsonResponse(
            {
                "success": False,
                "status": "error",
                "message": (
                    "Product images must be provided "
                    "inside a ZIP file."
                ),
            },
            status=400,
        )

    # ======================================================
    # FILE SIZES
    # ======================================================

    if excel_file.size > MAX_EXCEL_SIZE:
        return JsonResponse(
            {
                "success": False,
                "status": "error",
                "message": (
                    "The Excel file is too large. "
                    "Maximum size is 2 MB."
                ),
            },
            status=400,
        )

    if images_zip.size > MAX_ZIP_SIZE:
        return JsonResponse(
            {
                "success": False,
                "status": "error",
                "message": (
                    "The images ZIP is too large. "
                    "Maximum size is 10 MB."
                ),
            },
            status=400,
        )

    workbook = None
    images_archive = None

    try:
        # ==================================================
        # OPEN EXCEL
        # ==================================================

        try:
            workbook = load_workbook(
                filename=excel_file,
                data_only=True,
                read_only=True,
            )

        except Exception as error:
            return JsonResponse(
                {
                    "success": False,
                    "status": "error",
                    "message": (
                        "Unable to read the Excel file: "
                        f"{error}"
                    ),
                },
                status=400,
            )

        worksheet = workbook.active

        product_row_count = max(
            worksheet.max_row - 1,
            0,
        )

        if product_row_count == 0:
            return JsonResponse(
                {
                    "success": False,
                    "status": "error",
                    "message": (
                        "The Excel file contains no "
                        "product rows."
                    ),
                },
                status=400,
            )

        if (
            product_row_count
            > MAX_BULK_PRODUCTS
        ):
            return JsonResponse(
                {
                    "success": False,
                    "status": "error",
                    "message": (
                        f"Import a maximum of "
                        f"{MAX_BULK_PRODUCTS} products "
                        "at one time."
                    ),
                },
                status=400,
            )

        # ==================================================
        # OPEN ZIP
        # ==================================================

        try:
            images_zip.seek(0)

            images_archive = zipfile.ZipFile(
                images_zip,
                mode="r",
            )

        except zipfile.BadZipFile:
            return JsonResponse(
                {
                    "success": False,
                    "status": "error",
                    "message": (
                        "The selected images ZIP file "
                        "is invalid."
                    ),
                },
                status=400,
            )

        zip_image_map = build_zip_image_map(
            images_archive
        )

        if not zip_image_map:
            return JsonResponse(
                {
                    "success": False,
                    "status": "error",
                    "message": (
                        "The ZIP does not contain any "
                        "JPG, JPEG, PNG or WebP images."
                    ),
                },
                status=400,
            )

        imported_count = 0
        updated_count = 0
        skipped_count = 0
        errors = []

        # ==================================================
        # PROCESS EXCEL ROWS
        # ==================================================

        rows = worksheet.iter_rows(
            min_row=2,
            values_only=True,
        )

        for row_number, row in enumerate(
            rows,
            start=2,
        ):
            # Ignore completely empty rows.
            if not any(
                value not in (None, "")
                for value in row
            ):
                continue

            # Ensure at least ten values are available.
            row_values = (
                list(row)
                + [None] * 10
            )

            (
                category_name,
                product_name,
                price_value,
                offer_price_value,
                stock_value,
                short_description,
                long_description,
                image1,
                image2,
                image3,
            ) = row_values[:10]

            category_name = clean_excel_value(
                category_name
            )

            product_name = clean_excel_value(
                product_name
            )

            short_description = clean_excel_value(
                short_description
            )

            long_description = clean_excel_value(
                long_description
            )

            image_names = [
                normalize_image_filename(
                    image1
                ),
                normalize_image_filename(
                    image2
                ),
                normalize_image_filename(
                    image3
                ),
            ]

            image_names = [
                image_name
                for image_name in image_names
                if image_name
            ]

            # ----------------------------------------------
            # Required row fields
            # ----------------------------------------------

            if (
                not category_name
                or not product_name
            ):
                skipped_count += 1

                errors.append(
                    f"Row {row_number}: Category and "
                    "Product Name are required."
                )

                continue

            if not image_names:
                skipped_count += 1

                errors.append(
                    f"Row {row_number}: At least one "
                    "image is required."
                )

                continue

            row_uploaded_names = []

            try:
                # ------------------------------------------
                # Parse values
                # ------------------------------------------

                price = parse_decimal(
                    price_value,
                    "Price",
                    required=True,
                )

                offer_price = parse_decimal(
                    offer_price_value,
                    "Offer Price",
                    required=False,
                )

                stock_quantity = parse_stock(
                    stock_value
                )

                # ------------------------------------------
                # Validate all images before uploading
                # ------------------------------------------

                for image_name in image_names:
                    if (
                        image_name.lower()
                        not in zip_image_map
                    ):
                        raise ValueError(
                            f"Image '{image_name}' was "
                            "not found inside the ZIP."
                        )

                product_was_created = False
                old_images_list = []

                # ------------------------------------------
                # Database transaction for this row
                # ------------------------------------------

                with transaction.atomic():
                    category = (
                        Category.objects
                        .filter(
                            category_name__iexact=(
                                category_name
                            )
                        )
                        .first()
                    )

                    if category is None:
                        category = Category.objects.create(
                            category_name=category_name,
                            description="",
                            is_active=True,
                        )

                    product = (
                        Product.objects
                        .filter(
                            product_name__iexact=(
                                product_name
                            )
                        )
                        .first()
                    )

                    if product is None:
                        product = Product(
                            product_name=product_name
                        )

                        product_was_created = True

                    product.category = category
                    product.price = price

                    product.stock_quantity = (
                        stock_quantity
                    )

                    product.short_description = (
                        short_description
                    )

                    product.long_description = (
                        long_description
                    )

                    product.is_active = True

                    # Your model may contain offer_price
                    # or discount_price.
                    if hasattr(
                        product,
                        "offer_price",
                    ):
                        product.offer_price = (
                            offer_price
                        )

                    elif hasattr(
                        product,
                        "discount_price",
                    ):
                        product.discount_price = (
                            offer_price
                        )

                    product.save()

                    new_product_images = []

                    # --------------------------------------
                    # Upload images to Cloudinary
                    # --------------------------------------

                    for (
                        image_index,
                        image_name,
                    ) in enumerate(
                        image_names
                    ):
                        saved_name = (
                            upload_zip_product_image(
                                images_archive,
                                zip_image_map,
                                image_name,
                            )
                        )

                        row_uploaded_names.append(
                            saved_name
                        )

                        product_image = (
                            ProductImage.objects.create(
                                product=product,
                                image=saved_name,
                                is_primary=(
                                    image_index == 0
                                ),
                            )
                        )

                        new_product_images.append(
                            product_image
                        )

                    new_image_ids = [
                        product_image.id
                        for product_image
                        in new_product_images
                    ]

                    # --------------------------------------
                    # Replace old product image records
                    # --------------------------------------

                    old_images_queryset = (
                        ProductImage.objects
                        .filter(
                            product=product
                        )
                        .exclude(
                            id__in=new_image_ids
                        )
                    )

                    old_images_list = list(
                        old_images_queryset
                    )

                    old_images_queryset.delete()

                # ------------------------------------------
                # Delete old physical files after success
                # ------------------------------------------

                for old_image in old_images_list:
                    try:
                        old_image.image.delete(
                            save=False
                        )

                    except Exception:
                        logger.exception(
                            "Could not delete old "
                            "product image %s",
                            old_image.id,
                        )

                if product_was_created:
                    imported_count += 1

                else:
                    updated_count += 1

            except Exception as row_error:
                skipped_count += 1

                errors.append(
                    f"Row {row_number} "
                    f"({product_name or 'Unnamed Product'}): "
                    f"{row_error}"
                )

                # Remove Cloudinary files created by the
                # failed row.
                delete_saved_storage_files(
                    row_uploaded_names
                )

                logger.exception(
                    "Bulk import row %s failed",
                    row_number,
                )

        # ==================================================
        # FINAL RESPONSE
        # ==================================================

        message = (
            "Bulk import completed. "
            f"{imported_count} added, "
            f"{updated_count} updated, "
            f"{skipped_count} skipped."
        )

        return JsonResponse(
            {
                "success": True,
                "status": "success",
                "message": message,
                "reload_required": True,

                "summary": {
                    "added": imported_count,
                    "updated": updated_count,
                    "skipped": skipped_count,
                },

                "errors": errors[:30],
            }
        )

    finally:
        if images_archive is not None:
            images_archive.close()

        if workbook is not None:
            workbook.close()
            
# ==========================================================
# BULK PRODUCT IMPORT VIEW
# ==========================================================

@login_required
@role_permission_required("Can Manage Products")
@require_POST
def import_products(request):
    try:
        return process_product_bulk_import(request)

    except Exception as error:
        error_traceback = traceback.format_exc()

        logger.error(
            "BULK PRODUCT IMPORT FAILED\n%s",
            error_traceback
        )

        print(
            "BULK PRODUCT IMPORT FAILED",
            flush=True
        )

        print(
            error_traceback,
            flush=True
        )

        # Temporarily use status 200 so your current
        # JavaScript displays the real backend error.
        return JsonResponse({
            "success": False,
            "status": "error",
            "message": (
                f"{error.__class__.__name__}: {error}"
            ),
            "debug": error_traceback,
        }, status=200)

# ==========================================================
# COMMON BLOG QUERY HELPERS
# ==========================================================

def get_published_blog_posts():
    """
    Only posts currently visible on the customer website.
    """

    return (
        get_blog_posts_queryset()
        .filter(
            is_published=True,
            published_at__lte=timezone.now()
        )
    )
